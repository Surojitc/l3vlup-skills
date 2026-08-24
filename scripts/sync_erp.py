#!/usr/bin/env python3
"""
Equity risk premiums and country risk premiums, from Damodaran's published data.

Two workbooks, both updated by Damodaran himself:

  ctryprem.xlsx   sheet "ERPs by country"  — annual, each January
                  per-country Moody's rating, rating-based default spread,
                  total equity risk premium and country risk premium, plus a
                  sovereign rating-to-spread ladder in the right-hand columns

  ERPbymonth.xlsx sheet "Historical ERP"   — monthly
                  the implied ERP for the S&P 500, computed from the index level
                  and expected cash flows rather than from a historical average

We republish the figures with attribution and a link back, so the site can show
what the number currently is instead of telling the reader to go and find it.
Values are converted to percent; nothing is smoothed, rounded or interpolated.
"""

from __future__ import annotations

import io
import json
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

UA = {"User-Agent": "L3VLUP open skills contact@l3vlup.com"}

CTRYPREM = [
    "https://pages.stern.nyu.edu/~adamodar/pc/datasets/ctryprem.xlsx",
    "https://www.stern.nyu.edu/~adamodar/pc/datasets/ctryprem.xlsx",
]
ERP_MONTH = [
    "https://pages.stern.nyu.edu/~adamodar/pc/implprem/ERPbymonth.xlsx",
    "https://www.stern.nyu.edu/~adamodar/pc/implprem/ERPbymonth.xlsx",
]

HOME = "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datacurrent.html"
OUT = Path(__file__).resolve().parents[1] / "data" / "erp.auto.json"

#: Percent, to two decimals. Damodaran stores these as decimals.
def pct(v) -> float | None:
    if not isinstance(v, (int, float)):
        return None
    return round(float(v) * 100, 3)


def fetch(urls: list[str]) -> bytes:
    last = None
    for u in urls:
        try:
            req = urllib.request.Request(u, headers=UA)
            with urllib.request.urlopen(req, timeout=60) as r:
                return r.read()
        except Exception as e:  # try the mirror before giving up
            last = e
            print(f"  {u} -> {type(e).__name__}: {e}")
    raise RuntimeError(f"all mirrors failed: {last}")


def load(data: bytes, sheet: str):
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"sheet {sheet!r} not in {wb.sheetnames}")
    return list(wb[sheet].iter_rows(values_only=True))


def parse_countries(rows) -> tuple[list[dict], list[dict], str | None, float | None, float | None]:
    """
    The country table starts at whichever row has 'Country' in the first cell.
    Locating it by content rather than by index means the parse survives
    Damodaran adding or removing a note line above it, which he does most years.
    """
    header_i = next(
        (i for i, r in enumerate(rows)
         if r and isinstance(r[0], str) and r[0].strip().lower() == "country"),
        None,
    )
    if header_i is None:
        raise ValueError("could not find the 'Country' header row")

    # Two scalars Damodaran puts above the table: the S&P 500 implied ERP and
    # the US equity risk premium he is using as the mature-market base.
    sp500, us_erp, updated = None, None, None
    for r in rows[:header_i]:
        if not r:
            continue
        label = " ".join(str(c) for c in r if isinstance(c, str)).lower()
        nums = [c for c in r if isinstance(c, (int, float))]
        if "date of update" in label:
            d = next((c for c in r if isinstance(c, datetime)), None)
            if d:
                updated = d.date().isoformat()
        if "implied erp for s&p" in label and nums:
            sp500 = pct(nums[-1])
        elif "us equity risk premium" in label and nums:
            us_erp = pct(nums[-1])

    countries: list[dict] = []
    ladder: list[dict] = []
    for r in rows[header_i + 1:]:
        if not r:
            continue
        name = r[0]
        if isinstance(name, str) and name.strip():
            total = pct(r[4]) if len(r) > 4 else None
            if total is not None:
                countries.append({
                    "country": name.strip(),
                    "region": (r[1] or "").strip() if isinstance(r[1], str) else "",
                    "rating": (r[2] or "").strip() if isinstance(r[2], str) else "",
                    "defaultSpread": pct(r[3]) if len(r) > 3 else None,
                    "totalErp": total,
                    "crp": pct(r[5]) if len(r) > 5 else None,
                })
        # the sovereign rating ladder sits in the right-hand columns
        if len(r) > 10 and isinstance(r[9], str) and isinstance(r[10], (int, float)):
            rating = r[9].strip()
            if rating and rating.lower() != "rating":
                ladder.append({"rating": rating, "spreadBp": round(float(r[10]), 1)})

    return countries, ladder, updated, sp500, us_erp


def parse_implied(rows) -> list[dict]:
    """Monthly implied ERP. Column names shift year to year, so match on text."""
    header = rows[0]
    def col(*needles: str) -> int | None:
        for i, h in enumerate(header):
            if isinstance(h, str):
                t = h.strip().lower()
                if all(n in t for n in needles):
                    return i
        return None

    i_date = 0
    i_erp = col("erp", "t12m")
    i_tbond = col("t.bond")
    if i_erp is None:
        raise ValueError(f"no ERP column in {header}")

    out = []
    for r in rows[1:]:
        if not r or not isinstance(r[i_date], datetime):
            continue
        v = pct(r[i_erp]) if len(r) > i_erp else None
        if v is None:
            continue
        out.append({
            "d": r[i_date].date().isoformat(),
            "v": v,
            "tbond": pct(r[i_tbond]) if i_tbond is not None and len(r) > i_tbond else None,
        })
    return out


def main() -> None:
    print("ctryprem.xlsx")
    countries, ladder, updated, sp500, us_erp = parse_countries(
        load(fetch(CTRYPREM), "ERPs by country"))
    print(f"  {len(countries)} countries · {len(ladder)} rating rungs · updated {updated}")
    print(f"  S&P 500 implied ERP {sp500}% · US ERP {us_erp}%")

    print("ERPbymonth.xlsx")
    history = parse_implied(load(fetch(ERP_MONTH), "Historical ERP"))
    print(f"  {len(history)} monthly observations, latest {history[-1] if history else 'none'}")

    if not countries or not history:
        raise SystemExit("refusing to write a partial file")

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceUpdated": updated,
        "source": {
            "name": "Aswath Damodaran, NYU Stern",
            "url": HOME,
            "countryFile": CTRYPREM[0],
            "impliedFile": ERP_MONTH[0],
        },
        "sp500ImpliedErp": sp500,
        "usEquityRiskPremium": us_erp,
        "latestImplied": history[-1],
        # Ten years of monthly implied ERP is enough to show the range without
        # bloating the payload the site has to fetch.
        "impliedHistory": history[-120:],
        "ratingLadder": sorted(ladder, key=lambda x: x["spreadBp"]),
        "countries": sorted(countries, key=lambda c: c["country"]),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, separators=(",", ":")) + "\n")
    print(f"wrote {OUT} ({OUT.stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    main()
