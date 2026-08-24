#!/usr/bin/env python3
"""
Throwaway probe: report the shape of Damodaran's published datasets so a parser
can be written against what is actually there rather than against a guess.

Run from CI, where egress is open. Prints sheet names, header rows and a few
sample rows for each workbook.
"""
import io
import json
import urllib.request

UA = {"User-Agent": "L3VLUP open skills contact@l3vlup.com"}

TARGETS = {
    "ctryprem": [
        "https://pages.stern.nyu.edu/~adamodar/pc/datasets/ctryprem.xlsx",
        "https://www.stern.nyu.edu/~adamodar/pc/datasets/ctryprem.xlsx",
    ],
    "implied_erp_by_month": [
        "https://pages.stern.nyu.edu/~adamodar/pc/implprem/ERPbymonth.xlsx",
        "https://www.stern.nyu.edu/~adamodar/pc/implprem/ERPbymonth.xlsx",
    ],
    "ratings_spreads": [
        "https://pages.stern.nyu.edu/~adamodar/pc/datasets/ratings.xls",
        "https://pages.stern.nyu.edu/~adamodar/pc/ratings.xls",
    ],
    "betas_by_sector": [
        "https://pages.stern.nyu.edu/~adamodar/pc/datasets/betas.xls",
    ],
}


def fetch(urls):
    for u in urls:
        try:
            req = urllib.request.Request(u, headers=UA)
            with urllib.request.urlopen(req, timeout=45) as r:
                data = r.read()
            print(f"  OK  {u}  ({len(data)/1024:.0f} KB)")
            return u, data
        except Exception as e:
            print(f"  --  {u}  {type(e).__name__}: {e}")
    return None, None


def describe(name, data, url):
    import openpyxl
    print(f"\n{'='*72}\n{name}  <- {url}\n{'='*72}")
    if url.endswith(".xls"):
        print("  legacy .xls — openpyxl cannot read; needs xlrd or an .xlsx mirror")
        head = data[:8]
        print(f"  magic bytes: {head!r}")
        return
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        print(f"  load failed: {type(e).__name__}: {e}")
        return
    print(f"  sheets: {wb.sheetnames}")
    for sn in wb.sheetnames[:4]:
        ws = wb[sn]
        print(f"\n  --- sheet '{sn}' ({ws.max_row} rows x {ws.max_column} cols)")
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=14, max_col=12,
                                              values_only=True), start=1):
            cells = ["" if c is None else str(c)[:22] for c in row]
            if any(cells):
                print(f"    r{ri:<3} | " + " | ".join(cells))


def main():
    for name, urls in TARGETS.items():
        print(f"\n### {name}")
        url, data = fetch(urls)
        if data:
            describe(name, data, url)


if __name__ == "__main__":
    main()
