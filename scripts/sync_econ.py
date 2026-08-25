#!/usr/bin/env python3
"""
The economic backdrop a valuation sits on, from primary sources.

Four things a cost-of-capital screen needs and we did not collect:

  SOFR            New York Fed reference rates. The secured overnight rate that
                  replaced USD LIBOR, calculated from actual transaction volume
                  rather than submissions, and the base most floating-rate loans
                  are now quoted against.
  Short Treasuries  1-month, 3-month and 6-month bills off the Treasury's own
                  daily par yield curve. The short end the money market prices.
  Growth and inflation  IMF World Economic Outlook, which publishes history AND
                  forecasts. The forecast pair is what caps a terminal growth
                  assumption: nothing compounds faster than nominal GDP forever.
  Marginal tax rates  Damodaran's country table, sourced from the Tax Foundation.
                  Marginal, not effective, which is the rate a WACC tax shield
                  needs.

No API keys. Written against shapes a CI probe reported, not against guesses.
"""

from __future__ import annotations

import csv
import io
import json
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

UA = {"User-Agent": "L3VLUP open skills contact@l3vlup.com"}
OUT = Path(__file__).resolve().parents[1] / "data" / "econ.auto.json"

SOFR_URL = "https://markets.newyorkfed.org/api/rates/secured/sofr/last/30.json"
IMF_BASE = "https://www.imf.org/external/datamapper/api/v1"
TAX_URL = "https://pages.stern.nyu.edu/~adamodar/pc/datasets/countrytaxrates.xlsx"

#: ISO3 for the markets the site covers, mapped to display names.
COUNTRIES = {
    "USA": "United States", "GBR": "United Kingdom", "DEU": "Germany",
    "FRA": "France", "JPN": "Japan", "CHN": "China", "IND": "India",
    "BRA": "Brazil", "CAN": "Canada", "AUS": "Australia", "EA": "Euro Area",
}

def _norm(name: str) -> str:
    """Lowercase, drop punctuation and the filler words country lists disagree on."""
    t = name.lower()
    for junk in (" of great britain and northern ireland", " of america", " (the)",
                 ", the", "the ", "&", ".", ",", "'"):
        t = t.replace(junk, " ")
    return " ".join(t.split())


def match_tax(name: str, table: dict[str, float]) -> tuple[float | None, str | None]:
    """
    Damodaran's country spellings do not match the IMF's.

    The first live run returned None for both the United States and the United
    Kingdom, which is exactly the pair whose long-form names differ. Rather than
    guess an alias list, normalise both sides and fall back to a containment
    match, so a spelling we have not seen resolves instead of silently blanking.
    """
    if name in table:
        return table[name], name
    target = _norm(name)
    normalised = {_norm(k): (v, k) for k, v in table.items()}
    if target in normalised:
        return normalised[target]
    hits = [(v, k) for nk, (v, k) in normalised.items() if target in nk or nk in target]
    # Ambiguous is the same as unknown. One hit is a match; two is a guess.
    return hits[0] if len(hits) == 1 else (None, None)


def get(url: str, *, retries: int = 3) -> bytes:
    last = None
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=60) as r:
                return r.read()
        except Exception as e:
            last = e
            print(f"  attempt {attempt + 1}: {type(e).__name__}: {e}")
    raise RuntimeError(f"{url} failed: {last}")


def sofr() -> dict | None:
    """Latest SOFR print plus a short history, straight from the NY Fed."""
    rows = json.loads(get(SOFR_URL)).get("refRates", [])
    rows = [r for r in rows if r.get("type") == "SOFR" and isinstance(r.get("percentRate"), (int, float))]
    if not rows:
        return None
    rows.sort(key=lambda r: r["effectiveDate"])
    latest = rows[-1]
    return {
        "rate": latest["percentRate"],
        "asOf": latest["effectiveDate"],
        "volumeBn": latest.get("volumeInBillions"),
        "history": [{"d": r["effectiveDate"], "v": r["percentRate"]} for r in rows],
        "source": {"name": "Federal Reserve Bank of New York",
                   "url": "https://www.newyorkfed.org/markets/reference-rates/sofr"},
    }


def short_treasuries() -> dict | None:
    """
    The short end of the par yield curve. Same endpoint sync-macro.mjs already
    uses for the 2Y and 10Y, so the whole curve comes from one source.
    """
    year = datetime.now(timezone.utc).year
    url = (f"https://home.treasury.gov/resource-center/data-chart-center/interest-rates/"
           f"daily-treasury-rates.csv/{year}/all?type=daily_treasury_yield_curve"
           f"&field_tdr_date_value={year}&page&_format=csv")
    try:
        text = get(url).decode("utf-8-sig")
    except Exception as e:
        print(f"  treasury: {e}")
        return None
    rows = list(csv.DictReader(io.StringIO(text)))
    if not rows:
        return None
    # The file is newest-first; take the most recent row that has the tenors.
    wanted = {"1 Mo": "m1", "3 Mo": "m3", "6 Mo": "m6", "1 Yr": "y1"}
    for row in rows:
        out = {}
        for col, key in wanted.items():
            raw = (row.get(col) or "").strip()
            try:
                out[key] = float(raw)
            except ValueError:
                continue
        if out:
            out["asOf"] = row.get("Date")
            out["source"] = {
                "name": "U.S. Department of the Treasury",
                "url": "https://home.treasury.gov/resource-center/data-chart-center/interest-rates",
            }
            return out
    return None


def imf(indicator: str) -> dict[str, dict[str, float]]:
    """
    IMF World Economic Outlook via the Datamapper API.

    The country path segment is advisory — the endpoint returns every country
    regardless — so filter here rather than trusting the request to have done it.
    """
    j = json.loads(get(f"{IMF_BASE}/{indicator}"))
    vals = j.get("values", {}).get(indicator, {})
    out: dict[str, dict[str, float]] = {}
    for iso, series in vals.items():
        if iso not in COUNTRIES:
            continue
        clean = {y: float(v) for y, v in series.items()
                 if isinstance(v, (int, float)) and y.isdigit()}
        if clean:
            out[iso] = clean
    return out


def tax_rates() -> tuple[dict[str, float], str | None]:
    """Marginal corporate tax rates. Header row is found by content, not index."""
    wb = openpyxl.load_workbook(io.BytesIO(get(TAX_URL)), data_only=True, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    updated = None
    for r in rows[:6]:
        if r and isinstance(r[0], str) and "date updated" in r[0].lower():
            d = next((c for c in r if isinstance(c, datetime)), None)
            if d:
                updated = d.date().isoformat()
    header_i = next((i for i, r in enumerate(rows)
                     if r and isinstance(r[0], str) and r[0].strip().lower() == "country"), None)
    if header_i is None:
        raise ValueError("no 'Country' header in the tax workbook")
    out: dict[str, float] = {}
    for r in rows[header_i + 1:]:
        if r and isinstance(r[0], str) and isinstance(r[1], (int, float)):
            out[r[0].strip()] = round(float(r[1]) * 100, 2)
    return out, updated


def main() -> None:
    print("SOFR");            s = sofr()
    print("Treasury bills");  t = short_treasuries()
    print("IMF growth");      growth = imf("NGDP_RPCH")
    print("IMF inflation");   infl = imf("PCPIPCH")
    print("Tax rates");       taxes, tax_updated = tax_rates()

    this_year = str(datetime.now(timezone.utc).year)
    next_year = str(int(this_year) + 1)

    countries = []
    for iso, name in COUNTRIES.items():
        g, p = growth.get(iso, {}), infl.get(iso, {})
        if not g and not p:
            continue
        tax, tax_row = match_tax(name, taxes)
        if tax is None:
            print(f"  no tax rate matched for {name}")
        elif tax_row != name:
            print(f"  {name} -> tax row '{tax_row}' ({tax}%)")
        countries.append({
            "iso": iso,
            "country": name,
            "realGrowthForecast": g.get(next_year, g.get(this_year)),
            "realGrowthCurrent": g.get(this_year),
            "inflationForecast": p.get(next_year, p.get(this_year)),
            "inflationCurrent": p.get(this_year),
            # The ceiling a terminal growth assumption cannot exceed.
            "nominalGdpForecast": (
                round(g[next_year] + p[next_year], 2)
                if next_year in g and next_year in p else None
            ),
            "marginalTaxRate": tax,
            # The row the rate actually came from, so a wrong containment match is
            # visible on inspection rather than passing as fact. The UK figure in
            # particular has looked like a stale headline rate.
            "marginalTaxSourceRow": tax_row,
            "growthHistory": dict(sorted(g.items())[-12:]),
            "inflationHistory": dict(sorted(p.items())[-12:]),
        })

    if not countries or not s:
        raise SystemExit("refusing to write a partial file")

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sofr": s,
        "treasuryBills": t,
        "countries": sorted(countries, key=lambda c: c["country"]),
        "sources": {
            "sofr": "https://www.newyorkfed.org/markets/reference-rates/sofr",
            "treasury": "https://home.treasury.gov/resource-center/data-chart-center/interest-rates",
            "imf": "https://www.imf.org/external/datamapper/datasets/WEO",
            "tax": "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datacurrent.html",
            "taxOriginal": "Tax Foundation, via Damodaran",
            "taxUpdated": tax_updated,
        },
    }
    OUT.write_text(json.dumps(payload, separators=(",", ":")) + "\n")
    print(f"\nwrote {OUT} ({OUT.stat().st_size / 1024:.1f} KB)")
    print(f"  SOFR {s['rate']}% as at {s['asOf']}")
    if t:
        print(f"  bills 1m {t.get('m1')} · 3m {t.get('m3')} · 6m {t.get('m6')} as at {t.get('asOf')}")
    for c in countries[:4]:
        print(f"  {c['country']:<16} growth {c['realGrowthForecast']} · infl "
              f"{c['inflationForecast']} · nominal {c['nominalGdpForecast']} · tax {c['marginalTaxRate']}")


if __name__ == "__main__":
    main()
