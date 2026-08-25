#!/usr/bin/env python3
"""Throwaway: report the shape of the economic-data sources before writing parsers."""
import io, json, urllib.request

UA = {"User-Agent": "L3VLUP open skills contact@l3vlup.com"}

def get(url, label):
    try:
        with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=45) as r:
            data = r.read()
        print(f"\n### {label}\n  OK {url} ({len(data)/1024:.1f} KB)")
        return data
    except Exception as e:
        print(f"\n### {label}\n  -- {url} {type(e).__name__}: {e}")
        return None

# 1. SOFR — New York Fed reference rates API, free, no key
d = get("https://markets.newyorkfed.org/api/rates/secured/sofr/last/3.json", "SOFR (NY Fed)")
if d: print("  " + json.dumps(json.loads(d))[:700])

# 2. Treasury daily par yield curve — already used by sync-macro; want the 1m/3m/6m tenors
d = get("https://home.treasury.gov/resource-center/data-chart-center/interest-rates/daily-treasury-rates.csv"
        "?type=daily_treasury_yield_curve&field_tdr_date_value=2026", "Treasury par yield curve CSV")
if d:
    lines = d.decode("utf-8-sig").splitlines()
    print("  header:", lines[0][:300])
    print("  row1  :", lines[1][:300] if len(lines) > 1 else "none")

# 3. IMF Datamapper — real GDP growth and inflation, with forecasts, free, no key
for ind, label in (("NGDP_RPCH", "IMF real GDP growth"), ("PCPIPCH", "IMF inflation")):
    d = get(f"https://www.imf.org/external/datamapper/api/v1/{ind}/USA/GBR/DEU/JPN", label)
    if d:
        j = json.loads(d)
        print("  keys:", list(j.keys()))
        vals = j.get("values", {}).get(ind, {})
        for c, series in list(vals.items())[:2]:
            recent = {k: v for k, v in sorted(series.items())[-8:]}
            print(f"    {c}: {recent}")

# 4. Damodaran corporate tax rates by country
d = get("https://pages.stern.nyu.edu/~adamodar/pc/datasets/countrytaxrates.xlsx", "Damodaran country tax rates")
if d:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(d), data_only=True, read_only=True)
    print("  sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=8, values_only=True), 1):
        cells = ["" if c is None else str(c)[:24] for c in row]
        if any(cells): print(f"    r{ri:<3} | " + " | ".join(cells))
