"""
Banker house-style workbook writer.

Ported from the TWC model exporter (ClaudeSkills/scripts/fmp-model-export.py) so
L3VLUP skill outputs land in the same conventions an analyst already reads without
being told: blue is an input, black is a formula, green points at another sheet.

The one addition over the TWC original is PROVENANCE. Every figure can carry the
filing it came from; the writer collects those, numbers them, prints the marker
beside the row and emits a Sources sheet where each marker resolves to a tag,
period, form, accession number and a link to the filing on EDGAR. A number with
no provenance record is an unsourced number, and the workbook says so out loud.

Conventions (do not drift from these — they are the calibration contract):

  Colour       BLUE  = hard input / value pulled from a filing
               black = formula computed inside the workbook
               GREEN = link to another sheet
               NAVY  = headers, section labels, period row
               GREY  = notes, units, sub-rows

  Formats      D  $ in millions, negatives in parentheses
               P  percentage, one decimal
               X  multiple, one decimal, trailing "x"
               S  raw two-decimal (per-share)
               SH share count in millions
               N  plain integer

  Layout       column A gutter, B labels, C spacer, data from D
               gridlines off, landscape, fit-to-width, freeze panes at D7
               period row on row 6, header block on rows 2-5
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

BLUE = "FF0000FF"
NAVY = "FF264D82"
GREEN = "FF008000"
GREY = "FF808080"
RED = "FFC00000"

GUTTER, LBL, SPACER, DATA0 = 1, 2, 3, 4

CURRENCY_SYMBOL = {
    "USD": "$", "EUR": "€", "GBP": "£", "JPY": "¥",
    "CNY": "¥", "CHF": "CHF", "CAD": "C$", "AUD": "A$", "INR": "₹",
}

#: Masthead band. Every sheet opens with it so a page torn out of the workbook
#: still says where it came from.
BAND_FILL = PatternFill("solid", fgColor="FF0F172A")   # navy, the brand ground
BAND_ROWS = (1, 2, 3)
BAND_HEIGHTS = {1: 6, 2: 26, 3: 6}
#: Rows 1-3 are the band, 4-7 the header block, 8 the period row, 9 the first body row.
FIRST_BODY_ROW = 9
PERIOD_ROW = 8
ASSETS = Path(__file__).parent / "assets"

_THIN_BOTTOM = Border(bottom=Side(style="thin", color=NAVY))
_TOP_RULE = Border(top=Side(style="thin", color="FFBFBFBF"))


def fmt(marker: str, symbol: str = "$") -> str:
    """Excel number format for a house format marker."""
    if marker == "D":
        return f'"{symbol}"#,##0,,_);\\("{symbol}"#,##0,,\\)'
    if marker == "D0":  # already in millions, do not scale
        return f'"{symbol}"#,##0_);\\("{symbol}"#,##0\\)'
    if marker == "P":
        return '0.0%_);\\(0.0%\\)'
    if marker == "X":
        return '0.0"x"'
    if marker == "S":
        return '0.00_);\\(0.00\\)'
    if marker == "SH":
        return '#,##0,,_);\\(#,##0,,\\)'
    if marker == "N":
        return '#,##0'
    if marker == "BP":
        return '#,##0" bp"'
    return "General"


@dataclass(frozen=True)
class Source:
    """Where a number came from. Filings only for historic financials."""

    label: str                 # "AAPL FY2025 10-K"
    detail: str = ""           # "us-gaap:Revenues, FY2025, filed 2025-10-31"
    url: str = ""              # link to the filing on EDGAR
    accession: str = ""        # 0000320193-25-000073

    def key(self) -> str:
        return f"{self.label}|{self.detail}|{self.accession}"


@dataclass
class Fact:
    """A value plus the filing it came from."""

    value: float | int | str | None
    source: Source | None = None


def fact(value, source: Source | None = None) -> Fact:
    return Fact(value, source)


@dataclass
class _Row:
    label: str
    cells: list
    marker: str
    bold: bool
    italic: bool
    indent: bool
    colour: str | None
    rule: bool


class Sheet:
    """One tab. Rows are appended top to bottom; call finish() to lay out chrome."""

    def __init__(self, book: "Book", ws, title: str, periods: Sequence[tuple[str, bool]],
                 subtitle: str = "", units: str | None = None, link_to_summary: bool = True):
        self.book = book
        self.ws = ws
        self.title = title
        self.periods = list(periods)
        self.subtitle = subtitle
        self.units = units if units is not None else book.units
        self.link_to_summary = link_to_summary
        self._rows: list[_Row] = []
        self._r = FIRST_BODY_ROW  # band 1-3, header 4-7, period row 8

    # ---- content -------------------------------------------------------

    def section(self, label: str) -> "Sheet":
        c = self.ws.cell(self._r, LBL, label)
        c.font = Font(bold=True, color=NAVY, size=10)
        for k in range(len(self.periods) + 1):
            self.ws.cell(self._r, LBL + k if k == 0 else DATA0 + k - 1).border = _THIN_BOTTOM
        self._r += 1
        return self

    def blank(self, n: int = 1) -> "Sheet":
        self._r += n
        return self

    def row(self, label: str, cells: Iterable, marker: str = "D", *, bold: bool = False,
            italic: bool = False, indent: bool = False, formula: bool = False,
            link: bool = False, rule: bool = False) -> "Sheet":
        """
        cells — raw values (blue: pulled/hard input), Fact objects (blue + sourced),
        or strings beginning "=" (black: formula computed in the workbook).
        Pass formula=True to force black, link=True to force green.
        """
        cells = list(cells)
        lc = self.ws.cell(self._r, LBL, ("    " if indent else "") + label)
        lc.font = Font(bold=bold, italic=italic or indent,
                       color=(GREY if indent or italic else None), size=10)
        if rule:
            lc.border = _TOP_RULE

        for k, raw in enumerate(cells):
            src = None
            if isinstance(raw, Fact):
                src, raw = raw.source, raw.value
            cell = self.ws.cell(self._r, DATA0 + k)
            cell.number_format = fmt(marker, self.book.symbol)
            if rule:
                cell.border = _TOP_RULE
            if raw is None or raw == "":
                cell.value = "—"
                cell.font = Font(color=GREY, size=10)
                cell.alignment = Alignment(horizontal="right")
                continue
            cell.value = raw
            is_formula = formula or (isinstance(raw, str) and raw.startswith("="))
            colour = GREEN if link else (None if is_formula else BLUE)
            cell.font = Font(bold=bold, italic=italic, color=colour, size=10)
            if src is not None:
                self.book.note_source(src)

        # provenance marker column, one per row, right of the data
        srcs = [c.source for c in cells if isinstance(c, Fact) and c.source is not None]
        if srcs:
            ids = sorted({self.book.source_id(s) for s in srcs})
            mc = self.ws.cell(self._r, DATA0 + len(self.periods) + 1,
                              "".join(f"[{i}]" for i in ids))
            mc.font = Font(color=GREY, size=8)
            mc.alignment = Alignment(horizontal="left")

        self._r += 1
        return self

    def pct_row(self, label: str, formulas: Sequence[str]) -> "Sheet":
        """Grey italic margin/growth sub-row — always a formula, never a hard number."""
        return self.row(label, formulas, "P", italic=True, indent=True, formula=True)

    def note(self, text: str) -> "Sheet":
        c = self.ws.cell(self._r, LBL, text)
        c.font = Font(italic=True, color=GREY, size=8)
        self._r += 1
        return self

    def check(self, label: str, formulas: Sequence[str], marker: str = "D") -> "Sheet":
        """Balance / tie-out row: flags red when non-zero."""
        self.row(label, formulas, marker, italic=True, formula=True)
        r = self._r - 1
        first = get_column_letter(DATA0)
        last = get_column_letter(DATA0 + max(len(self.periods), 1) - 1)
        from openpyxl.formatting.rule import CellIsRule
        self.ws.conditional_formatting.add(
            f"{first}{r}:{last}{r}",
            CellIsRule(operator="notEqual", formula=["0"],
                       font=Font(bold=True, color="FFFFFFFF"),
                       fill=PatternFill("solid", fgColor=RED)),
        )
        return self

    # ---- chrome --------------------------------------------------------

    @property
    def cursor(self) -> int:
        """Row number the next write lands on."""
        return self._r

    def at(self, back: int = 1) -> int:
        """Row number of the row written `back` writes ago (1 = the last one)."""
        return self._r - back

    def col(self, index: int) -> str:
        """Column letter for period index (0-based) — for writing cross-sheet formulas."""
        return get_column_letter(DATA0 + index)

    def finish(self) -> "Sheet":
        ws, n = self.ws, len(self.periods)
        last_col_idx = DATA0 + max(n, 1) + 1

        _masthead(ws, last_col_idx)

        if self.link_to_summary and self.book.has_summary and self.title != self.book.summary_name:
            ws.cell(4, LBL, f"=+'{self.book.summary_name}'!B4").font = Font(
                bold=True, color=GREEN, size=11)
        else:
            ws.cell(4, LBL, self.book.company).font = Font(bold=True, color=NAVY, size=11)
        ws.cell(5, LBL, self.title).font = Font(bold=True, color=NAVY, size=12)
        if self.subtitle:
            ws.cell(6, LBL, self.subtitle).font = Font(italic=True, color=GREY, size=9)
        if self.units:
            ws.cell(7, LBL, self.units).font = Font(italic=True, color=GREY, size=9)

        if self.periods:
            ws.cell(7, DATA0, self.book.period_caption).font = Font(
                bold=True, color=NAVY, size=9)
            for k, (label, _is_est) in enumerate(self.periods):
                c = ws.cell(PERIOD_ROW, DATA0 + k, label)
                c.font = Font(bold=True, color=NAVY, size=10)
                c.alignment = Alignment(horizontal="right")
                c.border = _THIN_BOTTOM
            if self.book.sources:
                sc = ws.cell(PERIOD_ROW, DATA0 + n + 1, "Src")
                sc.font = Font(bold=True, color=GREY, size=8)

        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 1.7
        for k in range(max(n, 1)):
            ws.column_dimensions[get_column_letter(DATA0 + k)].width = 13
        ws.column_dimensions[get_column_letter(DATA0 + n)].width = 1.7
        ws.column_dimensions[get_column_letter(DATA0 + n + 1)].width = 10

        last_col = get_column_letter(last_col_idx)
        ws.print_area = f"A1:{last_col}{max(self._r, FIRST_BODY_ROW)}"
        ws.freeze_panes = f"D{FIRST_BODY_ROW}"
        ws.oddHeaderFooter = ws.oddFooter
        ws.oddFooter.left.text = "L3VLUP"
        ws.oddFooter.left.size = 8
        ws.oddFooter.left.color = "808080"
        ws.oddFooter.right.text = "Page &P of &N — see Cover for notices"
        ws.oddFooter.right.size = 8
        ws.oddFooter.right.color = "808080"
        return self


def _masthead(ws, last_col_idx: int, *, tall: bool = False) -> None:
    """
    The navy band with the wordmark, at the top of every sheet.

    A single tab is what actually gets printed, pasted into a deck or emailed on
    its own, so the identification has to live on the sheet rather than only on a
    cover page nobody exports.
    """
    heights = {1: 10, 2: 46, 3: 10} if tall else dict(BAND_HEIGHTS)
    for r in BAND_ROWS:
        ws.row_dimensions[r].height = heights[r]
        for c in range(1, last_col_idx + 1):
            ws.cell(r, c).fill = BAND_FILL

    logo = ASSETS / "l3vlup-logo-white.png"
    if logo.exists():
        try:
            from openpyxl.drawing.image import Image as XLImage

            img = XLImage(str(logo))
            # Read the aspect off the asset rather than hard-coding it, so the
            # band stays correct if the wordmark is ever redrawn.
            aspect = (img.width / img.height) if img.height else 5.36
            img.height = 32 if tall else 19
            img.width = round(img.height * aspect)
            img.anchor = "B2"
            ws.add_image(img)
            return
        except Exception:
            pass  # fall through to the text wordmark

    # Fallback when the raster is unavailable: the wordmark as styled text.
    c = ws.cell(2, LBL, "L3VLUP")
    c.font = Font(bold=True, color="FFFFFFFF", size=16 if tall else 12)
    c.alignment = Alignment(vertical="center")


class Book:
    """A workbook in the house style, with a Sources tab wired to every figure."""

    def __init__(self, company: str, *, currency: str = "USD",
                 units: str = "($ in millions, except per share)",
                 period_caption: str = "For Fiscal Year Ending",
                 summary_name: str = "Summary",
                 doc_title: str = "", skill: str = "", skill_url: str = ""):
        self.doc_title = doc_title or company
        self.skill = skill
        self.skill_url = skill_url
        self.generated = datetime.now(timezone.utc)
        self.company = company
        self.currency = currency
        self.symbol = CURRENCY_SYMBOL.get(currency, currency + " ")
        self.units = units
        self.period_caption = period_caption
        self.summary_name = summary_name
        self.has_summary = False
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self.sources: dict[str, tuple[int, Source]] = {}
        self._sheets: list[Sheet] = []

    def note_source(self, s: Source) -> int:
        return self.source_id(s)

    def source_id(self, s: Source) -> int:
        k = s.key()
        if k not in self.sources:
            self.sources[k] = (len(self.sources) + 1, s)
        return self.sources[k][0]

    def sheet(self, name: str, *, title: str | None = None,
              periods: Sequence[tuple[str, bool]] = (), subtitle: str = "",
              units: str | None = None) -> Sheet:
        ws = self.wb.create_sheet(name[:31])
        if name == self.summary_name:
            self.has_summary = True
        sh = Sheet(self, ws, title or name, periods, subtitle=subtitle, units=units)
        self._sheets.append(sh)
        return sh

    def write_cover_sheet(self, *, what_this_is: str = "",
                          source_note: str = "") -> None:
        """
        The front page: what the file is, how to read it, and the notices.

        It goes first in the tab order because the notices have to be seen before
        the numbers are used, not found afterwards by someone looking for them.
        """
        ws = self.wb.create_sheet("Cover", 0)
        LAST = 8
        _masthead(ws, LAST, tall=True)

        r = 5
        ws.cell(r, LBL, self.doc_title).font = Font(bold=True, color=NAVY, size=18)
        r += 1
        if self.skill:
            ws.cell(r, LBL, f"Produced by the L3VLUP {self.skill} skill").font = Font(
                italic=True, color=GREY, size=10)
            r += 1
        ws.cell(r, LBL, "Generated " + self.generated.strftime("%d %B %Y at %H:%M UTC")
                ).font = Font(italic=True, color=GREY, size=10)
        r += 2

        def block(title: str, lines: list[str], *, colour: str = NAVY,
                  body_colour: str | None = None, size: int = 9) -> None:
            nonlocal r
            c = ws.cell(r, LBL, title)
            c.font = Font(bold=True, color=colour, size=10)
            c.border = _THIN_BOTTOM
            for k in range(1, LAST):
                ws.cell(r, LBL + k).border = _THIN_BOTTOM
            r += 1
            for line in lines:
                cell = ws.cell(r, LBL, line)
                cell.font = Font(color=body_colour or GREY, size=size)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=r, start_column=LBL, end_row=r, end_column=LAST)
                ws.row_dimensions[r].height = max(14, 13 * (len(line) // 110 + 1))
                r += 1
            r += 1

        if what_this_is:
            block("What this is", [what_this_is])

        block("How to read it", [
            "BLUE  \u2014  a hard input, or a value pulled straight from a filing. Something a human can change.",
            "BLACK  \u2014  a formula computed inside this workbook. Do not overtype it.",
            "GREEN  \u2014  a link to another sheet in this workbook.",
            "RED FILL  \u2014  a tie-out that failed. Investigate before using the workbook.",
            "AN EM DASH  \u2014  the filer did not tag that line. It does not mean zero.",
            "Every [n] marker beside a row resolves on the Sources tab to a tag, a period, a form and an accession number.",
        ])

        block("Where the numbers come from", [
            source_note or
            "Historic financials are taken from SEC EDGAR XBRL company facts \u2014 the filer\u2019s own tagged "
            "figures as submitted. Not an aggregator, not a screener, not a language model\u2019s recollection. "
            "Share price and forward consensus are not filing data and are left as marked input cells rather "
            "than guessed.",
        ])

        block("Important notices", [
            "EDUCATIONAL USE ONLY. This workbook is produced by L3VLUP as a training and learning aid. It is "
            "not investment research, not a recommendation, and not an offer or solicitation to buy or sell any "
            "security or financial instrument.",

            "NOT INVESTMENT, LEGAL, TAX OR ACCOUNTING ADVICE. Nothing in this file constitutes advice of any "
            "kind and it takes no account of the objectives, financial situation or needs of any person. Obtain "
            "advice from a suitably qualified and regulated professional before acting on anything here.",

            "NO RELIANCE. The figures are assembled automatically from third-party filings and may be "
            "incomplete, mis-tagged at source, superseded by a later restatement, or wrong. Independently "
            "verify every figure against the underlying filing \u2014 which is why each one is linked \u2014 before "
            "relying on it for any purpose. Past performance is not a guide to future performance.",

            "NO WARRANTY. This workbook is provided \u201cas is\u201d and \u201cas available\u201d, without warranty of any "
            "kind, express or implied, including any warranty of accuracy, completeness, merchantability or "
            "fitness for a particular purpose.",

            "LIMITATION OF LIABILITY. To the fullest extent permitted by law, L3VLUP and its officers, employees "
            "and contributors accept no liability for any loss or damage \u2014 including any direct, indirect, "
            "incidental, consequential or economic loss, loss of profit, or loss of data \u2014 arising out of or in "
            "connection with the use of, or reliance on, this workbook or anything in it. Nothing in this notice "
            "limits or excludes any liability that cannot lawfully be limited or excluded.",

            "THIRD-PARTY DATA. Filing data originates from the U.S. Securities and Exchange Commission\u2019s EDGAR "
            "system and is in the public domain. The SEC does not endorse, sponsor or verify this workbook. "
            "Any figures you enter yourself remain governed by the terms of whichever data provider supplied "
            "them, and it is your responsibility to hold the necessary licence for them.",
        ], colour=RED, body_colour="FF333333")

        block("Copyright", [
            f"\u00a9 {self.generated.year} L3VLUP. All rights reserved.",
            "The structure, layout, formulas, wording and conventions of this workbook are the property of "
            "L3VLUP. It is licensed to the named recipient for their own personal study and professional work. "
            "You may not resell it, redistribute it, publish it, or use it to build or train a competing product "
            "or dataset, in whole or in part, without prior written permission.",
            "Underlying filing data is public domain and is not claimed as L3VLUP property.",
            "l3vlup.com" + (f"  \u00b7  {self.skill_url}" if self.skill_url else ""),
        ])

        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 22
        for k in range(3, LAST + 1):
            ws.column_dimensions[get_column_letter(k)].width = 14
        ws.print_area = f"A1:{get_column_letter(LAST)}{r}"

    def write_sources_sheet(self, *, unsourced_note: str = "") -> None:
        """
        The audit trail. Emitted last so it sees every figure written above it.
        This is the tab that separates a workbook you can defend from one you cannot.
        """
        ws = self.wb.create_sheet("Sources")
        _masthead(ws, 6)
        ws.cell(4, LBL, self.company).font = Font(bold=True, color=NAVY, size=11)
        ws.cell(5, LBL, "Sources and audit trail").font = Font(bold=True, color=NAVY, size=12)
        ws.cell(6, LBL, "Every marker in the [n] column of any tab resolves here. "
                        "Historic financials come from filings, never from an aggregator."
                ).font = Font(italic=True, color=GREY, size=9)
        if unsourced_note:
            ws.cell(7, LBL, unsourced_note).font = Font(italic=True, color=RED, size=9)

        heads = ["#", "Source", "Detail", "Accession", "Link"]
        for k, h in enumerate(heads):
            c = ws.cell(PERIOD_ROW, LBL + k, h)
            c.font = Font(bold=True, color=NAVY, size=10)
            c.border = _THIN_BOTTOM

        r = FIRST_BODY_ROW
        for _, (i, src) in sorted(self.sources.items(), key=lambda kv: kv[1][0]):
            ws.cell(r, LBL, i).font = Font(color=GREY, size=9)
            ws.cell(r, LBL + 1, src.label).font = Font(size=10)
            ws.cell(r, LBL + 2, src.detail).font = Font(color=GREY, size=9)
            ws.cell(r, LBL + 3, src.accession).font = Font(color=GREY, size=9)
            if src.url:
                c = ws.cell(r, LBL + 4, src.url)
                c.hyperlink = src.url
                c.font = Font(color="FF0563C1", underline="single", size=9)
            r += 1

        if not self.sources:
            ws.cell(FIRST_BODY_ROW, LBL + 1, "No sourced figures in this workbook."
                    ).font = Font(italic=True, color=RED, size=10)

        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 4
        ws.column_dimensions["C"].width = 34
        ws.column_dimensions["D"].width = 56
        ws.column_dimensions["E"].width = 24
        ws.column_dimensions["F"].width = 70
        ws.freeze_panes = f"C{FIRST_BODY_ROW}"

    def save(self, path, *, unsourced_note: str = "", what_this_is: str = "",
             source_note: str = "") -> None:
        """
        Sources first so it sees every figure written above it, then the cover,
        which is inserted at position 0 so the notices open with the file.
        """
        self.write_sources_sheet(unsourced_note=unsourced_note)
        self.write_cover_sheet(what_this_is=what_this_is, source_note=source_note)
        self.wb.save(path)
