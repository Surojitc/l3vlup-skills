#!/usr/bin/env python3
"""
Comps Set Builder — a defensible trading-comps sheet built from filings.

    python3 skills/comps-set-builder/build.py NVDA --peers AMD,AVGO,MRVL --out out/

Fundamentals (revenue, EBITDA, debt, cash, share count) come from each filer's own
XBRL tags, sourced to the accession number. Share price and forward consensus do
NOT come from filings and are therefore left as blue input cells for you to fill
from your own licensed feed. The workbook never invents them, and it never pretends
a screener number is a filing number.

Multiples are formulas. Change a price and every multiple, the median and the mean
move with it — which is the only way a comps sheet survives being questioned.

The Inclusion tab is the part interviewers actually push on: every name carries a
written reason it belongs in the set, and the names you rejected are listed with
the reason they were rejected. A comp set you cannot defend line by line is not a
comp set, it is a list.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from lib import financials  # noqa: E402
from lib.banker_xlsx import BLUE, Book, Fact  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

EST_FILL = PatternFill("solid", fgColor="FFFFF2CC")


def _ltm_or_fy(st: financials.Statements, key: str) -> Fact:
    return st.latest(key)


def build(target: str, peers: list[str], out_dir: Path, years: int = 3) -> Path:
    names = [target.upper()] + [p.upper() for p in peers if p.upper() != target.upper()]
    loaded: dict[str, financials.Statements] = {}
    failed: list[tuple[str, str]] = []
    for t in names:
        try:
            loaded[t] = financials.load(t, years=years)
        except Exception as exc:  # a ticker EDGAR does not know is a fact, not a crash
            failed.append((t, str(exc).split("\n")[0]))
    if not loaded:
        raise SystemExit("No tickers resolved on EDGAR — check the symbols.")

    cols = [(t, False) for t in loaded]
    n = len(cols)
    book = Book(f"{target.upper()} comparable companies",
                period_caption="Comparable companies")

    # ------------------------------------------------------------------ Comps
    s = book.sheet("Comps", title="Trading comparables",
                   periods=cols,
                   subtitle="Fundamentals from SEC filings · price and forward consensus "
                            "are your inputs, shaded")
    s.section("Market")
    s.row("Share price (input)", [None] * n, "S")
    price = s.at()
    for k in range(n):
        c = s.ws.cell(price, 4 + k)
        c.fill = EST_FILL
        c.font = Font(color=BLUE, size=10)
    s.row("Diluted shares outstanding", [loaded[t].latest("shares_diluted") for t in loaded], "SH")
    sh = s.at()
    s.row("Market capitalisation",
          [f'=IFERROR({s.col(k)}{price}*{s.col(k)}{sh},"")' for k in range(n)],
          bold=True, formula=True)
    mcap = s.at()
    s.blank()

    s.section("Capital structure (as filed)")
    s.row("Total debt", [
        Fact(
            None if all(loaded[t].latest(k2).value is None
                        for k2 in ("short_term_debt", "long_term_debt"))
            else sum(loaded[t].latest(k2).value for k2 in ("short_term_debt", "long_term_debt")
                     if loaded[t].latest(k2).value is not None),
            next((loaded[t].latest(k2).source for k2 in ("long_term_debt", "short_term_debt")
                  if loaded[t].latest(k2).source), None),
        ) for t in loaded
    ])
    debt = s.at()
    s.row("Cash and equivalents", [loaded[t].latest("cash") for t in loaded])
    cash = s.at()
    s.row("Short-term investments", [loaded[t].latest("short_term_investments") for t in loaded])
    sti = s.at()
    s.row("Net debt / (net cash)",
          [f'=IFERROR({s.col(k)}{debt}-{s.col(k)}{cash}-{s.col(k)}{sti},"")' for k in range(n)],
          bold=True, formula=True)
    nd = s.at()
    s.row("Enterprise value",
          [f'=IFERROR({s.col(k)}{mcap}+{s.col(k)}{nd},"")' for k in range(n)],
          bold=True, formula=True)
    ev = s.at()
    s.blank()

    s.section("Operating (latest fiscal year as filed)")
    s.row("Revenue", [loaded[t].latest("revenue") for t in loaded])
    rev = s.at()
    s.row("Gross profit", [loaded[t].latest("gross_profit") for t in loaded])
    gp = s.at()
    s.pct_row("% gross margin",
              [f'=IFERROR({s.col(k)}{gp}/{s.col(k)}{rev},"")' for k in range(n)])
    s.row("Operating income", [loaded[t].latest("operating_income") for t in loaded])
    ebit = s.at()
    s.row("Depreciation and amortisation", [loaded[t].latest("d_and_a") for t in loaded])
    da = s.at()
    s.row("EBITDA",
          [f'=IFERROR({s.col(k)}{ebit}+{s.col(k)}{da},"")' for k in range(n)],
          bold=True, formula=True)
    ebitda = s.at()
    s.pct_row("% EBITDA margin",
              [f'=IFERROR({s.col(k)}{ebitda}/{s.col(k)}{rev},"")' for k in range(n)])
    s.row("Net income", [loaded[t].latest("net_income") for t in loaded])
    s.row("Diluted EPS", [loaded[t].latest("eps_diluted") for t in loaded], "S")
    eps = s.at()
    s.blank()

    s.section("Trailing multiples")
    s.row("EV / Revenue", [f'=IFERROR({s.col(k)}{ev}/{s.col(k)}{rev},"")' for k in range(n)],
          "X", formula=True)
    evr = s.at()
    s.row("EV / EBITDA", [f'=IFERROR({s.col(k)}{ev}/{s.col(k)}{ebitda},"")' for k in range(n)],
          "X", formula=True)
    evebitda = s.at()
    s.row("P / E", [f'=IFERROR({s.col(k)}{price}/{s.col(k)}{eps},"")' for k in range(n)],
          "X", formula=True)
    pe = s.at()
    s.blank()

    s.section("Forward consensus (your input — not from filings)")
    s.row("NTM revenue (input)", [None] * n)
    ntm_rev = s.at()
    s.row("NTM EBITDA (input)", [None] * n)
    ntm_ebitda = s.at()
    s.row("NTM EPS (input)", [None] * n, "S")
    ntm_eps = s.at()
    for r in (ntm_rev, ntm_ebitda, ntm_eps):
        for k in range(n):
            c = s.ws.cell(r, 4 + k)
            c.fill = EST_FILL
            c.font = Font(color=BLUE, size=10)
    s.row("EV / NTM revenue",
          [f'=IFERROR({s.col(k)}{ev}/{s.col(k)}{ntm_rev},"")' for k in range(n)], "X", formula=True)
    fevr = s.at()
    s.row("EV / NTM EBITDA",
          [f'=IFERROR({s.col(k)}{ev}/{s.col(k)}{ntm_ebitda},"")' for k in range(n)], "X",
          formula=True)
    fevebitda = s.at()
    s.row("Forward P / E",
          [f'=IFERROR({s.col(k)}{price}/{s.col(k)}{ntm_eps},"")' for k in range(n)], "X",
          formula=True)
    fpe = s.at()
    s.blank()

    s.section("Set statistics")
    first, last = s.col(0), s.col(n - 1)
    for label, rows in (("Trailing", (evr, evebitda, pe)),
                        ("Forward", (fevr, fevebitda, fpe))):
        for name, r in zip(("EV / Revenue", "EV / EBITDA", "P / E"), rows):
            s.row(f"{label} {name} — median",
                  [f'=IFERROR(MEDIAN({first}{r}:{last}{r}),"")'] + [""] * (n - 1),
                  "X", formula=True, italic=True)
            s.row(f"{label} {name} — mean",
                  [f'=IFERROR(AVERAGE({first}{r}:{last}{r}),"")'] + [""] * (n - 1),
                  "X", formula=True, italic=True)
    s.blank()
    s.note("Shaded blue cells are yours to fill: price and consensus are not filing data, "
           "so this workbook will not guess them.")
    s.note("Multiples recompute from the cells above. Never hard-code a multiple.")
    if failed:
        s.note("Not resolved on EDGAR: " + "; ".join(f"{t} ({why})" for t, why in failed))
    s.finish()

    # -------------------------------------------------------------- Inclusion
    inc = book.sheet("Inclusion", title="Inclusion and rejection rationale",
                     periods=(), units="")
    inc.section("Why each name is in the set")
    inc.note("Fill the rationale column before you show this to anyone. Business-model "
             "similarity, size band, growth band, margin structure and end-market "
             "exposure are the five tests that survive scrutiny.")
    inc.blank()
    for t, st in loaded.items():
        inc.row(f"{t} — {st.name}", ["[rationale: business model / size / growth / margins / "
                                     "end market]"], "General")
    inc.blank()
    inc.section("Names considered and rejected")
    inc.note("An empty rejection list is a red flag. If you did not reject anyone, you did "
             "not build a set, you took a sector list.")
    for t, why in failed:
        inc.row(f"{t}", [f"Not resolved on EDGAR — {why}"], "General")
    for i in range(6):
        inc.row("[rejected ticker]", ["[reason for rejection]"], "General")
    inc.blank()
    inc.section("Adjustments made")
    inc.note("Note every adjustment here: stock-based compensation treatment, operating "
             "leases capitalised or not, non-recurring items excluded, fiscal-year "
             "calendarisation. An unstated adjustment is the fastest way to lose the room.")
    for i in range(4):
        inc.row("[adjustment]", ["[what and why]"], "General")
    inc.finish()

    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{target.upper()}_comps.xlsx"
    book.save(path, unsourced_note="Share price and forward consensus are user inputs and "
                                   "carry no filing provenance by design.")
    return path


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("ticker")
    ap.add_argument("--peers", default="", help="comma-separated peer tickers")
    ap.add_argument("--years", type=int, default=3)
    ap.add_argument("--out", default="out")
    a = ap.parse_args()
    peers = [p.strip() for p in a.peers.split(",") if p.strip()]
    print(f"wrote {build(a.ticker, peers, Path(a.out), a.years)}")


if __name__ == "__main__":
    main()
