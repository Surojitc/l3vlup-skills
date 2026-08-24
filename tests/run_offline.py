"""Build every skill workbook against fixture data — no network required."""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from tests.make_fixture import companyfacts  # noqa: E402
from lib import edgar  # noqa: E402

TICKERS = {"FIXA": "0000000001", "FIXB": "0000000002", "FIXC": "0000000003"}

edgar._ticker_map = lambda: TICKERS  # type: ignore[assignment]
edgar.cik_for = lambda t: TICKERS[t.upper()]  # type: ignore[assignment]
edgar.company_facts = lambda cik: companyfacts(  # type: ignore[assignment]
    {"0000000001": "Fixture Industries Inc",
     "0000000002": "Fixture Peer Holdings",
     "0000000003": "Fixture Rival Corp"}.get(cik, "Fixture Co"))
edgar.latest_filings = lambda cik, **k: [  # type: ignore[assignment]
    {"form": "10-K", "filed": "2026-02-14", "period": "2025-12-31",
     "accession": "0000000000-26-000001",
     "url": edgar.filing_url(cik, "0000000000-26-000001"), "primary": "form10k.htm"},
]

from lib import financials  # noqa: E402

# --- regression: comparatives must not shift the axis -----------------------
st = financials.load("FIXA", years=5)
assert st.labels == ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"], \
    f"fiscal labels wrong: {st.labels}"
rev = [f.value for f in st.annual("revenue")]
assert rev == sorted(rev), f"revenue not monotonic — axis is shuffled: {rev}"
assert abs(rev[-1] - 100e9) < 1, f"latest year is not the latest filing: {rev[-1]:,.0f}"
assert all(f.source for f in st.annual("revenue")), "a revenue year lost its source"
srcs = {f.source.label for f in st.annual("revenue")}
assert srcs == {f"FIXA {l} 10-K" for l in st.labels}, f"source labels drifted: {srcs}"
assert st.annual("d_and_a")[-1].value is not None, "D&A dropped off the axis"
from tests.make_fixture import STALE_TAG, LIVE_TAG  # noqa: E402

# --- regression: an abandoned tag must not win on preference order ----------
tags_used = {f.source.detail.split(":")[1].split(" ")[0]
             for f in st.annual("revenue") if f.source}
assert tags_used == {LIVE_TAG}, \
    f"stale tag {STALE_TAG} won on order instead of coverage: {tags_used}"
print("tag-switch regression: revenue resolved to", LIVE_TAG, "not", STALE_TAG)

print("axis regression: labels", st.labels, "· latest revenue",
      f"{rev[-1] / 1e9:.1f}bn · all sourced")

from skills.company_profile_shim import build as profile_build  # noqa: E402

out = ROOT / "out" / "fixture"
p = profile_build("FIXA", 5, out)
print("company-profile ->", p.relative_to(ROOT))

import openpyxl  # noqa: E402
from lib.banker_xlsx import BAND_FILL, FIRST_BODY_ROW  # noqa: E402

wb = openpyxl.load_workbook(p)
print("  sheets:", wb.sheetnames)
assert wb.sheetnames[0] == "Cover", "the notices must open with the file"
src = wb["Sources"]
n = sum(1 for r in range(FIRST_BODY_ROW, src.max_row + 1) if src.cell(r, 3).value)
print("  sourced figures indexed:", n)
assert n > 0, "no provenance recorded — the audit trail is empty"
ws = wb["Income Statement"]
assert ws.freeze_panes == f"D{FIRST_BODY_ROW}"
assert ws.sheet_view.showGridLines is False
print(f"  chrome ok (freeze D{FIRST_BODY_ROW}, gridlines off)")

# --- masthead and notices ---------------------------------------------------
for name in wb.sheetnames:
    sh = wb[name]
    band = [sh.cell(r, 2).fill.fgColor.rgb for r in (1, 2, 3)]
    assert all(b == BAND_FILL.fgColor.rgb for b in band), f"{name} has no masthead band"
    assert len(sh._images) >= 1 or sh.cell(2, 2).value == "L3VLUP", \
        f"{name} masthead carries no wordmark"
print("  masthead band + wordmark on all", len(wb.sheetnames), "tabs")

cover = wb["Cover"]
text = " ".join(str(cover.cell(r, 2).value or "") for r in range(1, cover.max_row + 1))
for phrase in ("EDUCATIONAL USE ONLY", "NOT INVESTMENT", "NO RELIANCE", "NO WARRANTY",
               "LIMITATION OF LIABILITY", "THIRD-PARTY DATA", "All rights reserved",
               "How to read it", "Where the numbers come from"):
    assert phrase in text, f"cover is missing: {phrase}"
print("  cover carries all notices, colour key and copyright")

from skills.comps_shim import build as comps_build  # noqa: E402

c = comps_build("FIXA", ["FIXB", "FIXC"], out)
print("comps-set-builder ->", c.relative_to(ROOT))
wbc = openpyxl.load_workbook(c)
print("  sheets:", wbc.sheetnames)
cs = wbc["Comps"]
labels = [cs.cell(r, 2).value for r in range(7, cs.max_row + 1) if cs.cell(r, 2).value]
for need in ("Enterprise value", "EV / EBITDA", "EV / NTM EBITDA"):
    assert any(need in str(x) for x in labels), f"missing row: {need}"
med = [cs.cell(r, 4).value for r in range(7, cs.max_row + 1)
       if str(cs.cell(r, 2).value or "").endswith("median")]
assert med and all(str(v).startswith("=IFERROR(MEDIAN") for v in med), "median rows not formulas"
print("  median rows are live formulas:", len(med))
srcc = wbc["Sources"]
print("  sourced figures indexed:", sum(1 for r in range(8, srcc.max_row + 1) if srcc.cell(r, 3).value))
print("\nALL OFFLINE CHECKS PASSED")
